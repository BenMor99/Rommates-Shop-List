# this program is designed to form an easy and better space for Expenses of dormroom

import openpyxl
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import colors
import specifier

redFill = PatternFill(start_color = 'FFFF0000',end_color = 'FFFF0000',fill_type = 'solid')
blueFill = PatternFill(start_color = '00B2EE',end_color = '00B2EE',fill_type = 'solid')
yellowFill = PatternFill(start_color = 'FFFF00',end_color = 'FFFF00',fill_type = 'solid')
greenFill = PatternFill(start_color = '11FF00',end_color = '11FF00',fill_type = 'solid')
whiteFill = PatternFill(start_color = 'FFFFFF',end_color = 'FFFFFF',fill_type = 'solid')
alphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']


def reset():

    wb = openpyxl.Workbook()
    ws = wb.active
    
    for i in range(1,19):
        if i <= 6:
            ws.cell(row = 1,column = i).fill = redFill
        elif i <= 12:
            ws.cell(row = 1 , column = i).fill = blueFill
        else:
            ws.cell(row = 1 , column = i).fill = yellowFill
            
    ws.cell(row=1,column=1).value = 'Behnam For Room'
    ws.cell(row=1,column=3).value = 'Behnam For Alireza'
    ws.cell(row=1,column=5).value = 'Behnam For Hamed'
    ws.cell(row=1,column=7).value = 'Alireza For Room'
    ws.cell(row=1,column=9).value = 'Alireza For Behnam'
    ws.cell(row=1,column=11).value = 'Alireza For Hamed'
    ws.cell(row=1,column=13).value = 'Hamed for Room'
    ws.cell(row=1,column=15).value = 'Hamed for Behnam'
    ws.cell(row=1,column=17).value = 'Hamed for Alireza'
    wb.save('Expenses.xlsx')
    specifier.spec()

def row_finder(c):
    wb = openpyxl.load_workbook('Expenses.xlsx',data_only=True)
    ws = wb.active
    counter = 1
    while ws.cell(row=counter,column=c).value is not None:
        counter += 1
    if counter == 1:
        return 0
    elif counter == 2:
        return 2
    else:
        return counter-1


class shop:

    def __init__(self, item, cost):
        self.item = item
        self.cost = cost

    def put(self,r,c):
        wb = openpyxl.load_workbook('Expenses.xlsx',data_only=True)
        ws = wb.active
        ws.cell(row = r,column = c).value = self.item
        ws.cell(row = r,column = c+1).value = int(self.cost)
        ws.cell(row = r,column = c).fill = whiteFill
        ws.cell(row = r,column = c+1).fill = whiteFill
        ws.cell(row = r+1,column = c).value = 'sum'
        ws.cell(row = r+1,column = c+1).value = '=SUM('+alphabet[c]+'2:'+alphabet[c]+str(r)+')'
        ws.cell(row = r+1,column = c).fill = greenFill
        ws.cell(row = r+1,column = c+1).fill = greenFill
        wb.save('Expenses.xlsx')
        specifier.spec()
        print('Successfully done!')

def data():
    item = input('What did you bought? ')
    cost = input('How much did that cost? ')
    who = input('for who? ')
    return [item , cost , who.lower()]

def location(your_name,for_who):
    if your_name == 'behnam':
        if for_who == 'room':
            return 1
        elif for_who == 'alireza':
            return 2
        elif for_who == 'hamed':
            return 3
        else:
            return 0

    elif your_name == 'alireza':
        if for_who == 'room':
            return 4
        elif for_who == 'behnam':
            return 5
        elif for_who == 'hamed':
            return 6
        else:
            return 0

    elif your_name == 'hamed':
        if for_who == 'room':
            return 7
        elif for_who == 'behnam':
            return 8
        elif for_who == 'alireza':
            return 9
        else:
            return 0

    else:
        return 0

    

def main():
    name = input('Who are you?')
    name = name.lower()
    if name=='behnam' or name=='hamed' or name=='alireza':
        d = data()
        c = location(name,d[2])
        if c == 0:
            print('Oops!\nIt seems there is no such name in your room!!!')
        else:
            s = shop(d[0],d[1])
            r = row_finder(2*c-1)
            if r == 0:
                print('Oops!\nI think you may want to reset first!!!')
            else:
                s.put(r,2*c-1)
     
    else:
        print('no names matched!')



